"""
app.py ─ 네이버 상가 수익률 분석기 (Streamlit 웹 앱)

실행:
    python -m streamlit run app.py

브라우저에서 http://localhost:8501 열기
"""

import io

import httpx
import pandas as pd
import streamlit as st
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill

from crawler_core import TokenExpiredError, crawl_region, fetch_regions

# ════════════════════════════════════════════════════════════════
# 페이지 설정
# ════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="네이버 상가 수익률 분석기",
    page_icon="🏢",
    layout="wide",
)

# ════════════════════════════════════════════════════════════════
# 세션 상태 초기화
# ════════════════════════════════════════════════════════════════

_DEFAULTS = {
    "sido_list":    [],   # [{"cortarNo": ..., "cortarName": ...}]
    "sigungu_list": [],
    "dong_list":    [],
    "results_df":   None,
}
for _k, _v in _DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v


# ════════════════════════════════════════════════════════════════
# 사이드바: 인증 정보
# ════════════════════════════════════════════════════════════════

def _load_sido():
    """JWT가 있으면 시/도 목록을 로드한다."""
    jwt    = st.session_state.get("jwt_input", "").strip()
    cookie = st.session_state.get("cookie_input", "").strip()
    # 하위 목록 초기화
    st.session_state.sido_list    = []
    st.session_state.sigungu_list = []
    st.session_state.dong_list    = []
    if jwt:
        try:
            with httpx.Client(verify=False, timeout=30) as c:
                st.session_state.sido_list = fetch_regions(c, "0000000000", jwt, cookie)
        except Exception:
            pass   # 경고는 메인 영역에서 표시


with st.sidebar:
    st.header("🔑 인증 정보")

    st.text_area(
        "JWT 토큰 (Bearer 뒤 값만 입력)",
        key="jwt_input",
        height=110,
        on_change=_load_sido,
        placeholder="eyJhbGci...",
    )
    st.text_area(
        "Cookie",
        key="cookie_input",
        height=80,
        placeholder="NID_SES=...",
    )

    # 수동 새로고침 버튼 (복붙 후 포커스를 잃지 않았을 때 대비)
    if st.button("🔄 지역 목록 새로고침", use_container_width=True):
        _load_sido()
        st.rerun()

    st.divider()
    st.caption(
        """
**JWT 발급 방법**
1. [new.land.naver.com](https://new.land.naver.com) 접속
2. F12 → Network 탭 열기
3. 아무 API 요청 클릭
4. Headers → `authorization: Bearer ...` 복사
5. `Bearer ` 이후 값만 붙여넣기

유효시간: 약 3시간
        """
    )

# ════════════════════════════════════════════════════════════════
# 메인: 타이틀
# ════════════════════════════════════════════════════════════════

st.title("🏢 네이버 상가 수익률 분석기")
st.caption("기보증금·월세 보유 매물만 필터링 | 수익률 = 월세 × 12 ÷ (매매가 − 기보증금)")

jwt    = st.session_state.get("jwt_input",    "").strip()
cookie = st.session_state.get("cookie_input", "").strip()

if not jwt:
    st.info("👈 좌측 사이드바에서 JWT 토큰을 입력하면 지역 목록이 로드됩니다.")
    st.stop()

if not st.session_state.sido_list:
    st.warning("JWT를 입력했지만 시/도 목록을 불러오지 못했습니다. 사이드바의 **지역 목록 새로고침** 버튼을 눌러보세요.")
    st.stop()

# ════════════════════════════════════════════════════════════════
# 지역 선택 (3단계 Cascading)
# ════════════════════════════════════════════════════════════════

st.subheader("📍 지역 선택")
col1, col2, col3 = st.columns(3)

# ── 1단계: 시/도 ────────────────────────────────────────────────────────────

def _on_sido_change():
    """시/도 변경 → 시/군/구 목록 재로드, 동 초기화"""
    sel    = st.session_state.get("sido_select")
    _jwt   = st.session_state.get("jwt_input", "").strip()
    _cookie = st.session_state.get("cookie_input", "").strip()
    st.session_state.sigungu_list = []
    st.session_state.dong_list    = []
    if sel and sel != "선택하세요":
        sido = next(
            (r for r in st.session_state.sido_list if r["cortarName"] == sel), None
        )
        if sido:
            try:
                with httpx.Client(verify=False, timeout=30) as c:
                    st.session_state.sigungu_list = fetch_regions(
                        c, sido["cortarNo"], _jwt, _cookie
                    )
            except Exception:
                pass

with col1:
    sido_names = ["선택하세요"] + [r["cortarName"] for r in st.session_state.sido_list]
    sido_sel = st.selectbox(
        "시/도",
        sido_names,
        key="sido_select",
        on_change=_on_sido_change,
    )

# ── 2단계: 시/군/구 ──────────────────────────────────────────────────────────

def _on_sigungu_change():
    """시/군/구 변경 → 동 목록 재로드"""
    sel     = st.session_state.get("sigungu_select")
    _jwt    = st.session_state.get("jwt_input", "").strip()
    _cookie = st.session_state.get("cookie_input", "").strip()
    st.session_state.dong_list = []
    if sel and sel != "전체":
        sigungu = next(
            (r for r in st.session_state.sigungu_list if r["cortarName"] == sel), None
        )
        if sigungu:
            try:
                with httpx.Client(verify=False, timeout=30) as c:
                    st.session_state.dong_list = fetch_regions(
                        c, sigungu["cortarNo"], _jwt, _cookie
                    )
            except Exception:
                pass

with col2:
    has_sigungu = bool(st.session_state.sigungu_list)
    sigungu_names = (
        ["전체"] + [r["cortarName"] for r in st.session_state.sigungu_list]
        if has_sigungu
        else ["─"]
    )
    sigungu_sel = st.selectbox(
        "시/군/구",
        sigungu_names,
        key="sigungu_select",
        on_change=_on_sigungu_change,
        disabled=not has_sigungu,
    )

# ── 3단계: 동 ───────────────────────────────────────────────────────────────

with col3:
    has_dong  = bool(st.session_state.dong_list)
    dong_names = (
        ["전체"] + [r["cortarName"] for r in st.session_state.dong_list]
        if has_dong
        else ["전체"]
    )
    dong_sel = st.selectbox(
        "동",
        dong_names,
        key="dong_select",
        disabled=not has_dong,
    )

# ════════════════════════════════════════════════════════════════
# 크롤링 대상 결정
# ════════════════════════════════════════════════════════════════

def _get_targets() -> list[dict]:
    """UI 선택 결과를 {cortarNo, label} 리스트로 변환."""
    if sido_sel == "선택하세요" or not st.session_state.sido_list:
        return []

    sido = next(
        (r for r in st.session_state.sido_list if r["cortarName"] == sido_sel), None
    )
    if not sido:
        return []

    # 시/군/구 전체 또는 목록 없음
    if sigungu_sel in ("전체", "─") or not st.session_state.sigungu_list:
        return [
            {
                "cortarNo": r["cortarNo"],
                "label":    f"{sido_sel} {r['cortarName']}",
            }
            for r in st.session_state.sigungu_list
        ]

    sigungu = next(
        (r for r in st.session_state.sigungu_list if r["cortarName"] == sigungu_sel),
        None,
    )
    if not sigungu:
        return []

    # 동 전체 또는 목록 없음
    if dong_sel == "전체" or not st.session_state.dong_list:
        return [
            {
                "cortarNo": sigungu["cortarNo"],
                "label":    f"{sido_sel} {sigungu_sel}",
            }
        ]

    dong = next(
        (r for r in st.session_state.dong_list if r["cortarName"] == dong_sel), None
    )
    if not dong:
        return []

    return [
        {
            "cortarNo": dong["cortarNo"],
            "label":    f"{sido_sel} {sigungu_sel} {dong_sel}",
        }
    ]


targets = _get_targets()

# ════════════════════════════════════════════════════════════════
# 크롤링 실행 버튼
# ════════════════════════════════════════════════════════════════

st.divider()
btn_col, info_col = st.columns([1, 3])

with btn_col:
    start_btn = st.button(
        "🚀 크롤링 시작",
        disabled=not targets,
        type="primary",
        use_container_width=True,
    )

with info_col:
    if not targets:
        st.warning("시/도부터 지역을 선택하세요.")
    elif len(targets) == 1:
        st.info(f"**{targets[0]['label']}** 크롤링 예정")
    else:
        st.info(f"**{sido_sel}** 전체 **{len(targets)}개** 시/군/구 크롤링 예정")

# ════════════════════════════════════════════════════════════════
# 크롤링 실행
# ════════════════════════════════════════════════════════════════

if start_btn:
    st.session_state.results_df = None

    status_box   = st.empty()
    region_bar   = st.progress(0.0, text="지역 준비 중...")
    detail_bar   = st.progress(0.0, text=" ")

    all_records: list[dict] = []
    total_regions = len(targets)
    token_expired = False

    with httpx.Client(verify=False, timeout=30) as client:
        for r_idx, region in enumerate(targets):
            cortar_no    = region["cortarNo"]
            region_label = region["label"]

            region_bar.progress(
                r_idx / total_regions,
                text=f"[{r_idx+1}/{total_regions}] {region_label}",
            )
            status_box.info(f"🔍 **{region_label}** 크롤링 중...")

            # 클로저 캡처를 위해 팩토리 함수 사용
            def _make_cb(bar):
                def _cb(done, total, msg):
                    bar.progress(done / max(total, 1), text=msg)
                return _cb

            try:
                records = crawl_region(
                    client,
                    cortar_no,
                    region_label,
                    jwt,
                    cookie,
                    progress_cb=_make_cb(detail_bar),
                )
                all_records.extend(records)
                status_box.success(f"✅ **{region_label}** : {len(records)}건 수집")

            except TokenExpiredError:
                status_box.error(
                    "🔴 **JWT 토큰이 만료**되었습니다. "
                    "사이드바에서 새 토큰을 입력한 뒤 다시 시작하세요."
                )
                token_expired = True
                break

            except Exception as e:
                status_box.warning(f"⚠️ {region_label} 오류: {e}")

            region_bar.progress(
                (r_idx + 1) / total_regions,
                text=f"[{r_idx+1}/{total_regions}] 완료",
            )

    # ── 결과 가공 ───────────────────────────────────────────────────────────
    if all_records and not token_expired:
        df = pd.DataFrame(all_records)

        # 기보증금 > 0 AND 월세 > 0 인 매물만
        mask = (df["기보증금(만원)"] > 0) & (df["월세(만원)"] > 0)
        df_filtered = df[mask].copy()

        # 수익률 내림차순 정렬
        df_sorted = (
            df_filtered
            .sort_values("수익률(%)", ascending=False)
            .reset_index(drop=True)
        )
        st.session_state.results_df = df_sorted

        detail_bar.empty()
        status_box.success(
            f"🎉 크롤링 완료!  "
            f"전체 **{len(df)}건** → 수익률 산출 가능 **{len(df_sorted)}건**"
        )

# ════════════════════════════════════════════════════════════════
# 결과 표시
# ════════════════════════════════════════════════════════════════

if st.session_state.results_df is not None:
    df = st.session_state.results_df

    if df.empty:
        st.warning("기보증금·월세 정보가 있는 매물이 없습니다.")
    else:
        st.subheader(f"📊 수익률 순위  ({len(df)}건)")
        st.caption("🔴 6% 이상  ·  🟠 4~6%  ·  🟢 4% 미만")

        # ── 표시용 컬럼 가공 ────────────────────────────────────────────────
        display_df = df.copy()

        def _yield_emoji(v):
            if v is None or pd.isna(v):
                return "-"
            emoji = "🔴" if v >= 6 else "🟠" if v >= 4 else "🟢"
            return f"{emoji} {v:.2f}%"

        display_df.insert(0, "수익률", display_df["수익률(%)"].apply(_yield_emoji))

        show_cols = [
            "수익률", "소재지", "매매대금(만원)", "기보증금(만원)", "월세(만원)",
            "계약면적(㎡)", "해당층", "총층", "방향", "매물특징", "중개사",
            "상세정보", "매물번호",
        ]

        st.dataframe(
            display_df[show_cols],
            use_container_width=True,
            height=620,
            column_config={
                "상세정보":       st.column_config.LinkColumn("상세정보"),
                "매매대금(만원)": st.column_config.NumberColumn(format="%d 만원"),
                "기보증금(만원)": st.column_config.NumberColumn(format="%d 만원"),
                "월세(만원)":     st.column_config.NumberColumn(format="%d 만원"),
                "수익률":         st.column_config.TextColumn("수익률(연)"),
            },
        )

        # ── 다운로드 ────────────────────────────────────────────────────────
        st.divider()
        dcol1, dcol2 = st.columns(2)

        # CSV
        with dcol1:
            csv_bytes = df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
            st.download_button(
                "⬇️ CSV 다운로드",
                csv_bytes,
                file_name="상가수익률.csv",
                mime="text/csv",
                use_container_width=True,
            )

        # Excel (openpyxl 조건부 서식 + 자동 너비)
        with dcol2:
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="수익률분석")
                ws = writer.sheets["수익률분석"]

                # 헤더 스타일
                header_fill = PatternFill(fill_type="solid", fgColor="2F5496")
                for cell in ws[1]:
                    cell.font      = Font(bold=True, color="FFFFFF")
                    cell.fill      = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # 열 너비 자동 조정
                for col in ws.columns:
                    max_len = max(len(str(c.value or "")) for c in col) + 4
                    ws.column_dimensions[col[0].column_letter].width = min(max_len, 42)

                # 첫 행 고정 + 자동 필터
                ws.freeze_panes  = "A2"
                ws.auto_filter.ref = ws.dimensions

                # 수익률(%) 컬럼 조건부 서식 (빨강 ≥6%, 주황 4~6%)
                yield_col_idx    = list(df.columns).index("수익률(%)") + 1
                yield_col_letter = ws.cell(1, yield_col_idx).column_letter
                last_row         = len(df) + 1
                yield_range      = f"{yield_col_letter}2:{yield_col_letter}{last_row}"

                ws.conditional_formatting.add(
                    yield_range,
                    CellIsRule(
                        operator="greaterThanOrEqual",
                        formula=["6"],
                        fill=PatternFill(fill_type="solid", fgColor="FFB3B3"),
                        font=Font(bold=True),
                    ),
                )
                ws.conditional_formatting.add(
                    yield_range,
                    CellIsRule(
                        operator="between",
                        formula=["4", "5.9999"],
                        fill=PatternFill(fill_type="solid", fgColor="FFE0B2"),
                    ),
                )

            st.download_button(
                "⬇️ Excel 다운로드",
                buf.getvalue(),
                file_name="상가수익률.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
